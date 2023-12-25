import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries, is_cell_empty
from .utils import BaseTableDetector


class TableDetector(BaseTableDetector):
    def find_table_end(self, openpyxl_ws, start_row, start_col):
        ws_range = openpyxl_ws.calculate_dimension(force=True)
        _, ws_max_row, _, ws_max_col = range_boundaries(ws_range)
        max_row = start_row
        max_col = start_col
        empty_row_count = 0
        empty_col_count = 0

        # Horizontal Expansion
        for j in range(start_col, ws_max_col + 1):
            column_empty = True
            for i in range(start_row, ws_max_row + 1):
                if not is_cell_empty(openpyxl_ws.cell(row=i, column=j)):
                    column_empty = False
                    max_row = max(max_row, i)  # Update max_row based on this column's data
                    break

            if column_empty:
                empty_col_count += 1
                if empty_col_count > 1:
                    break
            else:
                empty_col_count = 0
                max_col = j

        # Vertical Expansion
        for i in range(start_row, max_row + 1):
            row_empty = True
            for j in range(start_col, max_col + 1):
                if not is_cell_empty(openpyxl_ws.cell(row=i, column=j)):
                    row_empty = False
                    break

            if row_empty:
                empty_row_count += 1
                if empty_row_count > 1:
                    break
            else:
                empty_row_count = 0
                max_row = i

        return max_row, max_col

    def get_table_ranges(self, openpyxl_ws, **kwargs):
        tables = []
        visited = set()

        for row_idx, row in enumerate(openpyxl_ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_coord = f"{get_column_letter(col_idx)}{row_idx}"

                if cell_coord in visited or is_cell_empty(cell):
                    continue

                end_row, end_col = self.find_table_end(openpyxl_ws, row_idx, col_idx)
                table_range = f"{get_column_letter(col_idx)}{row_idx}:{get_column_letter(end_col)}{end_row}"
                tables.append(table_range)

                # Mark cells as visited
                for r in range(row_idx, end_row + 1):
                    for c in range(col_idx, end_col + 1):
                        visited.add(f"{get_column_letter(c)}{r}")

        return tables
